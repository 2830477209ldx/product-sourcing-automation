"""Shopify 导入表.xlsx exporter — matches the 75-column template format."""

from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl
from loguru import logger

from src.models.product import Product


class ShopifyExcelExporter:
    """Export products into the Shopify import Excel format (导入表.xlsx)."""

    # Column letters → header names (matching the template)
    HEADERS: dict[str, str] = {
        "A": "ID", "B": "Handle", "C": "Command", "D": "Title",
        "E": "Body HTML", "F": "Vendor", "G": "Type", "H": "Tags",
        "I": "Tags Command", "J": "Created At", "K": "Updated At",
        "L": "Status", "M": "Published", "N": "Published At",
        "O": "Published Scope", "P": "Template Suffix", "Q": "Gift Card",
        "R": "URL", "S": "Total Inventory Qty", "T": "Row #", "U": "Top Row",
        "V": "Category: ID", "W": "Category: Name", "X": "Category",
        "Y": "Custom Collections", "Z": "Smart Collections",
        "AA": "Image Type", "AB": "Image Src", "AC": "Image Command",
        "AD": "Image Position", "AE": "Image Width", "AF": "Image Height",
        "AG": "Image Alt Text",
        "AH": "Variant Inventory Item ID", "AI": "Variant ID", "AJ": "Variant Command",
        "AK": "Option1 Name", "AL": "Option1 Value",
        "AM": "Option2 Name", "AN": "Option2 Value",
        "AO": "Option3 Name", "AP": "Option3 Value",
        "AQ": "Variant Position", "AR": "Variant SKU", "AS": "Variant Barcode",
        "AT": "Variant Image", "AU": "Variant Weight", "AV": "Variant Weight Unit",
        "AW": "Variant Price", "AX": "Variant Compare At Price",
        "AY": "Variant Taxable", "AZ": "Variant Tax Code",
        "BA": "Variant Inventory Tracker", "BB": "Variant Inventory Policy",
        "BC": "Variant Fulfillment Service", "BD": "Variant Requires Shipping",
        "BE": "Variant Shipping Profile", "BF": "Variant Inventory Qty",
        "BG": "Variant Inventory Adjust",
        "BH": "Included / International", "BI": "Price / International",
        "BJ": "Compare At Price / International",
        "BK": "Included / United States", "BL": "Price / United States",
        "BM": "Compare At Price / United States",
        "BT": "Metafield: custom.description [rich_text_field]",
        "BU": "Metafield: custom.inspiration [rich_text_field]",
        "BV": "Metafield: custom.highlights [rich_text_field]",
        "BW": "Metafield: custom.notices [rich_text_field]",
    }

    COL_KEYS = [
        "A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K",
        "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U",
        "V", "W", "X", "Y", "Z",
        "AA", "AB", "AC", "AD", "AE", "AF", "AG",
        "AH", "AI", "AJ", "AK", "AL", "AM", "AN", "AO", "AP",
        "AQ", "AR", "AS", "AT", "AU", "AV", "AW", "AX", "AY", "AZ",
        "BA", "BB", "BC", "BD", "BE", "BF", "BG",
        "BH", "BI", "BJ", "BK", "BL", "BM",
        "BT", "BU", "BV", "BW",
    ]

    def __init__(self, vendor: str = "") -> None:
        self.vendor = vendor

    def export(
        self,
        products: list[dict[str, Any]],
        image_map: dict[str, list[str]] | None = None,
        metafields_map: dict[str, dict[str, str]] | None = None,
        output_path: str | Path = "data/exports/import_table.xlsx",
    ) -> Path:
        """Export products to Shopify import Excel.

        Args:
            products: List of product dicts with keys: handle, title, body_html,
                      vendor, tags, skus (list of dicts with name, price, compare_price,
                      weight_grams, image, inventory).
            image_map: {handle: [image_paths]} for Image Src columns.
            metafields_map: {handle: {description, inspiration, highlights, notices}}.
            output_path: Output .xlsx file path.

        Returns:
            Path to the generated Excel file.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet2"  # Match template sheet name

        # Write headers
        for col_key in self.COL_KEYS:
            if col_key in self.HEADERS:
                ws[f"{col_key}1"] = self.HEADERS[col_key]

        row_idx = 2
        for prod in products:
            handle = prod.get("handle", "")
            title = prod.get("title", "")
            body_html = prod.get("body_html", "")
            vendor = prod.get("vendor", self.vendor)
            tags = prod.get("tags", "")
            skus = prod.get("skus", [])
            if not isinstance(skus, list):
                skus = []

            imgs = (image_map or {}).get(handle, [])
            mf = (metafields_map or {}).get(handle, {})

            # Build product-level defaults
            defaults = {
                "B": handle, "D": title, "E": body_html,
                "F": vendor, "H": tags,
                "I": "REPLACE", "L": "Active", "M": "TRUE",
                "O": "global", "Q": "FALSE", "AJ": "MERGE",
                "AK": "Title" if not skus else "款式",
                "BT": mf.get("description", ""),
                "BU": mf.get("inspiration", ""),
                "BV": mf.get("highlights", ""),
                "BW": mf.get("notices", ""),
            }

            if not skus:
                # Single product without variants
                self._write_row(ws, row_idx, defaults)
                # Write images in the same row
                for i, img in enumerate(imgs[:20]):
                    if i == 0:
                        ws[f"AB{row_idx}"] = img
                    else:
                        # Additional images need extra rows (Shopify convention)
                        img_row = row_idx + i
                        self._write_cell(ws, img_row, "AB", img)
                row_idx += max(1, len(imgs))
            else:
                # Multi-variant: first row = product info, subsequent rows = variants
                self._write_row(ws, row_idx, defaults)
                # Images on first row
                for i, img in enumerate(imgs):
                    col = self._image_col(i)
                    if col:
                        ws[f"{col}{row_idx}"] = img
                row_idx += 1

                for vi, sku in enumerate(skus):
                    sku_name = sku.get("name", "")
                    sku_price = sku.get("price", "")
                    sku_compare = sku.get("compare_price", "")
                    sku_weight = sku.get("weight_grams", 0)
                    sku_image = sku.get("image", "")
                    sku_inventory = sku.get("inventory", 0)

                    variant = {
                        "B": handle,
                        "AJ": "MERGE",
                        "AK": "款式",
                        "AL": sku_name,
                        "AQ": vi + 1,
                        "AW": self._clean_price(sku_price),
                        "AX": self._clean_price(sku_compare),
                        "AU": sku_weight,
                        "AV": "g",
                        "AT": sku_image,
                        "BF": sku_inventory,
                        "BG": 0,
                        "AY": "TRUE",
                        "BA": "shopify",
                        "BB": "deny",
                        "BC": "manual",
                        "BD": "TRUE",
                    }
                    self._write_row(ws, row_idx, variant)
                    row_idx += 1

        output = Path(output_path)
        output.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(output))
        logger.success(f"Exported {len(products)} products to {output}")
        return output

    def _write_row(self, ws, row_idx: int, data: dict[str, Any]) -> None:
        for col, val in data.items():
            if val is not None and val != "":
                ws[f"{col}{row_idx}"] = val

    def _write_cell(self, ws, row_idx: int, col: str, val: Any) -> None:
        if val is not None and val != "":
            ws[f"{col}{row_idx}"] = val

    @staticmethod
    def _image_col(index: int) -> str | None:
        """AB=Image Src (index 0). Additional images go in the same column on new rows."""
        if index == 0:
            return "AB"
        return None  # Additional images are placed on separate rows

    @staticmethod
    def _clean_price(val: Any) -> str:
        """Normalize price to numeric string."""
        if val is None or val == "":
            return ""
        s = str(val).strip().replace("¥", "").replace("￥", "").replace("$", "").replace(",", "")
        try:
            return str(float(s))
        except ValueError:
            return ""


def export_products_to_xlsx(
    products: list[Product],
    image_paths: dict[str, list[str]] | None = None,
    metafields: dict[str, dict[str, str]] | None = None,
    output_path: str = "data/exports/import_table.xlsx",
) -> Path:
    """Convenience function: Product objects → Shopify import Excel.

    Args:
        products: List of Product domain objects.
        image_paths: {product_id: [webp_image_paths]}.
        metafields: {product_id: {description, inspiration, highlights, notices}}.
        output_path: Output .xlsx path.

    Returns:
        Path to generated file.
    """
    prod_dicts = []
    for p in products:
        pid = p.id or ""
        prod_dicts.append({
            "handle": p.make_handle(),
            "title": p.title_en or p.title_cn,
            "body_html": p.optimized_description or "",
            "vendor": "",
            "tags": ", ".join(p.tags) if p.tags else "",
            "skus": p.sku_prices if isinstance(p.sku_prices, list) else [],
        })
    return ShopifyExcelExporter().export(
        prod_dicts,
        image_map=image_paths,
        metafields_map=metafields,
        output_path=output_path,
    )
